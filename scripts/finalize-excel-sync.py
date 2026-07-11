# -*- coding: utf-8 -*-
"""全面核对 Excel vs DB，修正名称/图片/上架，并写 ensure 脚本所需数据。"""
from __future__ import annotations

import json
import re
import shutil
import sqlite3
import ssl
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public" / "products"
DBS = [ROOT / "dev.db", ROOT / "prisma" / "data" / "initial.db"]
XLSX = Path(r"c:\Users\33092\Desktop\网站数据.xlsx")

CTX = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode = ssl.CERT_NONE

# Excel 全名 → 站内 slug
SLUG_ALIASES = {
    "强化包 极巨攻防 BOX（简中）": "极巨攻防-box-简中",
    "强化包 璀璨反击 BOX（简中）": "璀璨反击-box-简中",
    "强化包 怒炎灼天 BOX（简中）": "怒炎灼天-box-简中",
    "强化包 终末炎舞 BOX（简中）": "终末炎舞-box-简中",
    "强化包 暗影夺辉 BOX（简中）": "暗影夺辉-box-简中",
    "强化包 胜象星引 BOX（简中）": "胜象星引-box-简中",
    "强化包 第一弹 BOX（简中）": "sm-强化包-第一弹-box-简中",
    "强化包 第二弹 BOX（简中）": "sm-强化包-第二弹-box-简中",
    "补充包 强化扩张包 朱紫ex BOX（简中）": "朱-紫ex-強化拡張パック-box-简中",
    # 朱紫肥盒：Excel「补充包 X 肥盒」= 站内主规格「X BOX」
    "补充包 奇迹启程 肥盒 BOX（简中）": "奇迹启程-box-简中",
    "补充包 无畏太晶 肥盒 BOX（简中）": "无畏太晶-box-简中",
    "补充包 嘉奖回合 肥盒 BOX（简中）": "嘉奖回合-box-简中",
    "补充包 黑晶炽诚 肥盒 BOX（简中）": "黑晶炽焰-box-简中",
    "补充包 真实玄虚 肥盒 BOX（简中）": "真实玄虚-box-简中",
    "补充包 利刃猛醒 肥盒 BOX（简中）": "剑刃觉醒-box-简中",
    "补充包 璀璨诡幻 肥盒 BOX（简中）": "璀璨诡幻-box-简中",
    "补充包 星彩晶璃 肥盒 BOX（简中）": "星彩晶璃-box-简中",
    "补充包 共逐荣光 肥盒 BOX（简中）": "共逐荣光-box-简中",
    "补充包 奇迹启程 瘦盒 BOX（简中）": "奇迹启程-slim-box-简中",
    "补充包 无畏太晶 瘦盒 BOX（简中）": "无畏太晶-slim-box-简中",
    "补充包 嘉奖回合 瘦盒 BOX（简中）": "嘉奖回合-slim-box-简中",
    "补充包 黑晶炽诚 瘦盒 BOX（简中）": "黑晶炽焰-slim-box-简中",
    "补充包 真实玄虚 瘦盒 BOX（简中）": "真实玄虚-slim-box-简中",
    "补充包 利刃猛醒 瘦盒 BOX（简中）": "剑刃觉醒-slim-box-简中",
    "补充包 璀璨诡幻 瘦盒 BOX（简中）": "璀璨诡幻-slim-box-简中",
    "补充包 星彩晶璃 瘦盒 BOX（简中）": "星彩晶璃-slim-box-简中",
    "补充包 共逐荣光 瘦盒 BOX（简中）": "共逐荣光-slim-box-简中",
}

# 名称规范化（站内展示名）
NAME_FIXES = {
    "151-コレクタ-ズセット-简中": ("151 收藏家套装（简中）", "151-collectors-set-简中"),
    "ポケモンカ-ド151-box-简中": ("宝可梦卡 151 BOX（简中）", "pokemon-card-151-box-简中"),
    "火影忍者-コレクタブルカ-ド-スタ-タ-セット": ("火影忍者 收藏卡入门套装", "naruto-collectible-starter-set"),
    "スノ-ハザ-ド-クレイバ-スト-ダブルパック-简中": ("雪危险+泥土猛击 双包（简中）", "sv2-double-pack-简中"),
}

# 黑炎霸主用简中官方图（无畏太晶/同世代 chase 不如 csv 黑炎）
IMAGE_BY_SLUG = {
    "朱-紫-黒炎の支配者-box-简中": "/products/sv3-box.png",  # 保留；若有更好图再换
}

OFFICIAL_IMGS = {
    "30th-deluxe-box.png": [
        "https://www.pokemon.cn/assets/img/products/2026/02/260220_30th_CGWL_650_488.png",
        "https://image.pokemon.com.cn/wp-content/uploads/2026/02/260220_30th_CGWL_650_488.png",
    ],
    "lillie-battle-box.png": [
        "https://www.pokemon.cn/assets/img/products/2026/05/260515_CSV10_LLY_650_488.png",
        "https://image.pokemon.com.cn/wp-content/uploads/2026/05/260515_CSV10_LLY_650_488.png",
    ],
    "n-battle-box.png": [
        "https://www.pokemon.cn/assets/img/products/2026/05/260515_CSV10_N_650_488.png",
        "https://image.pokemon.com.cn/wp-content/uploads/2026/05/260515_CSV10_N_650_488.png",
    ],
}


def now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")


def download(url: str, dest: Path) -> bool:
    try:
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Referer": "https://www.pokemon.cn/",
                "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
            },
        )
        with urllib.request.urlopen(req, timeout=45, context=CTX) as resp:
            data = resp.read()
        if data[:1] == b"<" or len(data) < 2000:
            print(f"  bad {url[:60]}… ({len(data)}b)")
            return False
        dest.write_bytes(data)
        print(f"  OK {dest.name} from {url.split('/')[-1]} ({len(data)}b)")
        return True
    except Exception as e:
        print(f"  fail {url[:60]}… {e}")
        return False


def ensure_official_images() -> None:
    print("=== 尝试官网真图 ===")
    for fn, urls in OFFICIAL_IMGS.items():
        dest = PUBLIC / fn
        # 若文件过大且是礼盒占位（火箭队/happy），仍尝试官网
        for url in urls:
            if download(url, dest):
                break
        else:
            print(f"  keep existing {fn} size={dest.stat().st_size if dest.exists() else 0}")


def load_excel_priced():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    rows = []
    for sheet in ("宝可梦", "海贼王"):
        ws = wb[sheet]
        for r in range(5, ws.max_row + 1):
            name = ws.cell(r, 1).value
            series = ws.cell(r, 2).value
            box = ws.cell(r, 3).value
            status = ws.cell(r, 4).value
            stock = ws.cell(r, 5).value
            price = ws.cell(r, 7).value
            if not name or not series:
                continue
            name = str(name).strip()
            if "官方名称" in name or "待官方" in name:
                continue
            if price in (None, ""):
                continue
            rows.append(
                {
                    "sheet": sheet,
                    "name": name,
                    "series": str(series).strip(),
                    "box": str(box or "").strip(),
                    "status": str(status or "").strip(),
                    "stock": int(stock or 0),
                    "price": int(float(price)),
                }
            )
    return rows


def find_product(db: sqlite3.Connection, name: str):
    p = db.execute(
        "SELECT id,name,slug,stock,priceJpy,status,images FROM Product WHERE name=?",
        (name,),
    ).fetchone()
    if p:
        return p
    if name in SLUG_ALIASES:
        p = db.execute(
            "SELECT id,name,slug,stock,priceJpy,status,images FROM Product WHERE slug=?",
            (SLUG_ALIASES[name],),
        ).fetchone()
        if p:
            return p
    alt = name.replace("补充包 ", "").replace("强化包 ", "")
    if alt != name:
        p = db.execute(
            "SELECT id,name,slug,stock,priceJpy,status,images FROM Product WHERE name=?",
            (alt,),
        ).fetchone()
        if p:
            return p
    # 模糊：去掉 BOX 前后缀差异
    core = re.sub(r"^(补充包|强化包)\s*", "", name)
    core = core.replace("强化扩张包 ", "").replace(" BOX（简中）", "").replace("（简中）", "")
    hits = db.execute(
        "SELECT id,name,slug,stock,priceJpy,status,images FROM Product WHERE name LIKE ? AND name LIKE '%简中%'",
        (f"%{core}%",),
    ).fetchall()
    if len(hits) == 1:
        return hits[0]
    # 优先主 BOX（不含散包/原箱）
    mains = [h for h in hits if "散包" not in h["name"] and "原箱" not in h["name"]]
    if len(mains) == 1:
        return mains[0]
    return None


def sync_db(db_path: Path, excel_rows) -> dict:
    if not db_path.exists():
        return {"path": str(db_path), "skip": True}
    db = sqlite3.connect(str(db_path))
    db.row_factory = sqlite3.Row
    updated = 0
    missing = []
    for e in excel_rows:
        p = find_product(db, e["name"])
        if not p:
            missing.append(e["name"])
            continue
        is_pre = 1 if str(e.get("status") or "").strip() == "预售" else 0
        db.execute(
            "UPDATE Product SET priceJpy=?, stock=?, status='上架', isPreorder=?, updatedAt=? WHERE id=?",
            (e["price"], e["stock"], is_pre, now(), p["id"]),
        )
        updated += 1

    # 名称/slug 修正
    name_fixed = 0
    for old_slug, (new_name, new_slug) in NAME_FIXES.items():
        row = db.execute("SELECT id,slug FROM Product WHERE slug=?", (old_slug,)).fetchone()
        if not row:
            # 也可能已是新 slug
            continue
        # 避免 slug 冲突
        conflict = db.execute("SELECT id FROM Product WHERE slug=? AND id!=?", (new_slug, row["id"])).fetchone()
        if conflict:
            print(f"  skip slug fix {old_slug} -> {new_slug} (conflict)")
            continue
        db.execute(
            "UPDATE Product SET name=?, slug=?, updatedAt=? WHERE id=?",
            (new_name, new_slug, now(), row["id"]),
        )
        name_fixed += 1
        print(f"  rename {old_slug} -> {new_slug}")

    # 全部上架
    draft = db.execute("UPDATE Product SET status='上架' WHERE status!='上架'").rowcount
    db.commit()

    # 复查缺图
    broken = []
    for r in db.execute("SELECT name,images FROM Product"):
        imgs = [x.strip() for x in (r[1] or "").replace("[", "").replace("]", "").replace('"', "").split(",") if x.strip()]
        if not imgs:
            broken.append((r[0], "EMPTY"))
            continue
        fp = ROOT / "public" / imgs[0].lstrip("/")
        if not fp.exists():
            broken.append((r[0], imgs[0]))

    total = db.execute("SELECT COUNT(*) FROM Product").fetchone()[0]
    listed = db.execute("SELECT COUNT(*) FROM Product WHERE status='上架'").fetchone()[0]
    db.close()
    return {
        "path": db_path.name,
        "updated": updated,
        "missing": missing,
        "name_fixed": name_fixed,
        "draft_fixed": draft,
        "broken": broken,
        "total": total,
        "listed": listed,
    }


def main():
    ensure_official_images()
    excel_rows = load_excel_priced()
    print(f"\n=== Excel 有价行 {len(excel_rows)} ===")
    for dbp in DBS:
        print(f"\n=== sync {dbp.name} ===")
        res = sync_db(dbp, excel_rows)
        print(json.dumps({k: v for k, v in res.items() if k != "missing"}, ensure_ascii=False, indent=2))
        if res.get("missing"):
            print("仍未匹配:")
            for m in res["missing"]:
                print(" ", m)
        if res.get("broken"):
            print("缺图:")
            for b in res["broken"]:
                print(" ", b)


if __name__ == "__main__":
    main()
