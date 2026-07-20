# -*- coding: utf-8 -*-
"""
按桌面「网站数据.xlsx」全量上架：
- 宝可梦 → 海贼王表格顺序写入 catalogSort
- 全部 status=上架
- 规格（瘦/肥/散包/原箱）拆成独立 SKU
- 图片优先表格路径，再按系列映射 fat/slim / onepiece-official，并保留已有顶卡
"""
from __future__ import annotations

import re
import sqlite3
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
_XLSX_CANDIDATES = [
    Path(r"c:\Users\33092\Desktop\网站数据.xlsx"),
    ROOT / "prisma" / "data" / "网站数据.xlsx",
    ROOT / "网站数据.xlsx",
]
XLSX = next((p for p in _XLSX_CANDIDATES if p.exists()), _XLSX_CANDIDATES[0])


def resolve_dbs() -> list[Path]:
    """优先写 DATABASE_URL（Railway Volume），并同步本地快照。"""
    import os

    out: list[Path] = []
    url = (os.environ.get("DATABASE_URL") or "").strip()
    if url.startswith("file:"):
        p = Path(url[5:])
        if not p.is_absolute():
            p = (ROOT / p).resolve()
        out.append(p)
    for p in (ROOT / "dev.db", ROOT / "prisma" / "data" / "initial.db"):
        if p.resolve() not in {x.resolve() for x in out}:
            out.append(p)
    return out


DBS = resolve_dbs()
PUBLIC = ROOT / "public"

# 站内中文展示名（表格仍为英文时覆盖）
EBC_DISPLAY = {
    "EBC-02": "EBC-02 动画25周年纪念合集",
    "EBC-03": "EBC-03 女英雄合集",
    "EBC-04": "EBC-04 艾格赫德危机",
}

CSV_TITLE = {
    "共逐荣光": "共逐荣光",
    "星彩晶璃": "星彩晶璃",
    "璀璨诡幻": "璀璨诡幻",
    "利刃猛醒": "剑刃觉醒",
    "剑刃觉醒": "剑刃觉醒",
    "真实玄虚": "真实玄虚",
    "黑晶炽诚": "黑晶炽焰",
    "黑晶炽焰": "黑晶炽焰",
    "嘉奖回合": "嘉奖回合",
    "无畏太晶": "无畏太晶",
    "奇迹启程": "奇迹启程",
    "太晶盛聚": "太晶盛聚",
    "亘古开来": "亘古开来",
}

OPC_SLUG_BASE = {
    "01": "opc-01-romance-dawn",
    "02": "opc-02-paramount-war",
    "03": "opc-03-pillars-of-strength",
    "04": "opc-04-kingdoms-of-intrigue",
    "05": "opc-05-awakening-new-era",
    "06": "opc-06-wings-of-the-captain",
    "07": "opc-07-500-years-future",
    "08": "opc-08-two-legends",
    "09": "opc-09-emperors-in-the-new-world",
    "10": "opc-10-royal-blood",
    "11": "opc-11-a-fist-of-divine-speed",
    "12": "opc-12-official",
    "13": "opc-13-official",
    "14": "opc-14-cyan-sea-seven",
    "15": "opc-15-adventure-gods-island",
    "16": "opc-16-official",
}


def now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")


def public_exists(rel: str | None) -> bool:
    if not rel or not str(rel).startswith("/"):
        return False
    return (PUBLIC / str(rel).lstrip("/")).exists()


def pick_image(candidates: list[str], fallback: str = "/products/placeholder.svg") -> str:
    for c in candidates:
        if c and public_exists(c):
            return c
    for c in candidates:
        if c:
            return c
    return fallback


def csv_code(series: str | None) -> str | None:
    if not series:
        return None
    m = re.search(r"CSV\s*(\d+(?:\.\d+)?)c", str(series), re.I)
    return f"CSV{m.group(1).replace('.', '')}C".replace("CSV95C", "CSV9.5C") if m else None


def csv_title_from_name(name: str) -> str | None:
    for k, v in CSV_TITLE.items():
        if k in name:
            return v
    m = re.search(r"补充包\s+(.+?)\s+(瘦盒|肥盒)", name)
    if m:
        return CSV_TITLE.get(m.group(1).strip(), m.group(1).strip())
    return None


def normalize_pokemon_box(box: str | None, spec: str | None) -> str | None:
    b = (box or "").strip()
    s = (spec or "").strip()
    if b in ("瘦盒", "肥盒"):
        if s == "散包":
            return "瘦散包" if b == "瘦盒" else "肥散包"
        if s == "原箱":
            return "瘦原箱" if b == "瘦盒" else "肥原箱"
        return b
    if "补充包(肥盒)" in b or b == "补充包(肥盒)":
        # 剑盾等仅肥盒轨：用原盒/散包/原箱，避免被「肥盒作规格附属」从列表隐藏
        if s == "散包":
            return "散包"
        if s == "原箱":
            return "原箱"
        return "原盒"
    if "强化包(瘦盒)" in b:
        if s == "散包":
            return "瘦散包"
        if s == "原箱":
            return "瘦原箱"
        return "瘦盒"
    if b == "宝石包":
        if s == "散包":
            return "散包"
        if s == "原箱":
            return "原箱"
        return "宝石包"
    if b in ("礼盒", "专属礼盒", "大礼盒", "构筑套装", "对战套装", "周边", "强化扩张包"):
        return b if s in ("", "原盒", None) or s == "原盒" else b
    if s in ("原盒", "散包", "原箱") and b in ("原盒", ""):
        return s
    return b or s or None


def format_sku_name(name: str, box_type: str) -> str:
    """表格大纲行常复用父名，按规格改写展示名。"""
    n = name.strip()
    swaps = {
        "瘦散包": [("瘦盒 BOX", "瘦散包"), ("瘦盒", "瘦散包")],
        "肥散包": [("肥盒 BOX", "肥散包"), ("肥盒", "肥散包")],
        "瘦原箱": [("瘦盒 BOX", "瘦原箱"), ("瘦盒", "瘦原箱")],
        "肥原箱": [("肥盒 BOX", "肥原箱"), ("肥盒", "肥原箱")],
        "散包": [("原盒", "散包")],
        "原箱": [("原盒", "原箱")],
    }
    for old, new in swaps.get(box_type, []):
        if old in n:
            return n.replace(old, new, 1)
    return n


def pokemon_slug(
    title: str | None,
    box_type: str,
    series: str | None,
    excel_slug: str | None,
    name: str,
) -> str:
    # 剑盾「补充包(肥盒)」映射后的原盒/散包/原箱：必须按规格拆 slug
    if box_type in ("原盒", "散包", "原箱") and (
        (series and ("剑" in series or "CS" in series))
        or "补充包" in name
        or "强化包" in name
    ):
        base = (title or re.sub(r"[^\w\u4e00-\u9fff-]+", "-", name.strip())[:40]).strip("-")
        suffix = {"原盒": "box", "散包": "pack", "原箱": "case"}[box_type]
        return f"{base}-{suffix}-简中"

    # 礼盒/周边等：绝不能走扩充包 slug（否则会占掉 pokemon-151-*-slim-box）
    if box_type not in SV_BOX:
        if excel_slug and str(excel_slug).strip():
            return str(excel_slug).strip()
        base = re.sub(r"[^\w\u4e00-\u9fff-]+", "-", name.strip())[:80]
        return f"{base}-简中" if not base.endswith("简中") else base

    # 151 四弹扩充包
    m151 = re.search(r"收集啦151\s*(旅|望|惊|聚)", f"{series or ''} {name}")
    if m151 and "硬币" not in name and "展示" not in name and "收藏家" not in name:
        key = {"旅": "lv", "望": "wang", "惊": "jing", "聚": "ju"}[m151.group(1)]
        fmt = {
            "瘦盒": "slim-box",
            "肥盒": "fat-box",
            "瘦散包": "slim-pack",
            "肥散包": "fat-pack",
            "瘦原箱": "slim-case",
            "肥原箱": "fat-case",
        }.get(box_type, "slim-box")
        return f"pokemon-151-{key}-{fmt}-cn"

    if title and box_type in SV_BOX:
        if box_type == "瘦盒":
            return f"{title}-slim-box-简中"
        if box_type == "肥盒":
            return f"{title}-box-简中"
        if box_type == "瘦散包":
            return f"{title}-slim-pack-简中"
        if box_type == "肥散包":
            return f"{title}-fat-pack-简中"
        if box_type == "瘦原箱":
            return f"{title}-slim-case-简中"
        if box_type == "肥原箱":
            return f"{title}-fat-case-简中"

    if excel_slug and str(excel_slug).strip():
        slug = str(excel_slug).strip()
        # 修正错误重复 slug
        if box_type.endswith("散包") and "-pack-" not in slug and not slug.endswith("-pack"):
            slug = re.sub(r"-box-简中$", "-pack-简中", slug)
            slug = re.sub(r"-box$", "-pack", slug)
            if slug == str(excel_slug).strip():
                slug = re.sub(r"-slim-box-简中$", "-slim-pack-简中", slug)
                slug = re.sub(r"-fat-box-简中$", "-fat-pack-简中", slug)
        if box_type.endswith("原箱") and "-case-" not in slug and not slug.endswith("-case"):
            slug = re.sub(r"-box-简中$", "-case-简中", slug)
            slug = re.sub(r"-pack-简中$", "-case-简中", slug)
            if slug == str(excel_slug).strip() or slug.endswith("-pack-简中"):
                slug = re.sub(r"-slim-box-简中$", "-slim-case-简中", slug)
                slug = re.sub(r"-fat-box-简中$", "-fat-case-简中", slug)
                slug = re.sub(r"-slim-pack-简中$", "-slim-case-简中", slug)
                slug = re.sub(r"-fat-pack-简中$", "-fat-case-简中", slug)
        return slug

    base = re.sub(r"[^\w\u4e00-\u9fff-]+", "-", (title or series or "item").strip())
    return f"{base}-{box_type}-简中"


SV_BOX = {"瘦盒", "肥盒", "瘦散包", "肥散包", "瘦原箱", "肥原箱"}


# 礼盒/套装等非扩充包：避免误挂同系列 CSV 瘦包面
GIFT_NAME_IMAGE = [
    ("N对战", "/products/n-battle-box.png"),
    ("嗨皮组合", "/products/happy-combo.png"),
    ("火箭队", "/products/rocket-gang-box.png"),
    ("莉莉艾", "/products/lillie-battle-box.png"),
    ("伊布礼盒", "/products/eevee-gift-box.png"),
    ("多龙巴鲁托", "/products/master-deck-hydreigon.png"),
    ("猛雷鼓", "/products/master-deck-thunderdrum.png"),
    ("赛富豪", "/products/master-deck-ceruledge.png"),
    ("最初的伙伴展示套装", "/products/151-box.png"),
    ("硬币收藏艺术展示框", "/products/151-box.png"),
]


def pokemon_images(box_type: str, series: str | None, excel_img: str | None, name: str) -> str:
    for key, path in GIFT_NAME_IMAGE:
        if key in name and public_exists(path):
            return path

    code = csv_code(series)
    if code:
        key = code.lower().replace("csv", "csv")
        # CSV9.5C → csv95? 站内多用 csv9c / 自定义
        num = re.search(r"(\d+(?:\.\d+)?)", code)
        stem = f"csv{num.group(1).replace('.', '')}c" if num else key
        if "9.5" in code or "95" in stem:
            stem = "csv95c" if public_exists(f"/products/csv95c-slim.png") else "csv9c"
        fat = f"/products/{stem}-fat.png"
        slim = f"/products/{stem}-slim.png"
        box = f"/products/{stem}-box.png"
        prefer = fat if box_type.startswith("肥") else slim
        cover = pick_image([prefer, excel_img or "", box, slim, fat])
        return cover

    tops_151 = ",".join(
        f"/products/topcards/151c-{i}.webp"
        for i in range(1, 6)
        if public_exists(f"/products/topcards/151c-{i}.webp")
    )

    m151 = re.search(r"收集啦151\s*(旅|望|惊|聚)", f"{series or ''} {name}")
    if m151:
        key = {"旅": "lv", "望": "wang", "惊": "jing", "聚": "ju"}[m151.group(1)]
        kind = "fat" if box_type.startswith("肥") else "slim"
        # 硬币套装等礼盒优先用盒面图
        if "硬币" in name or box_type in ("礼盒", "周边"):
            cover = pick_image(
                [f"/products/151c-{key}-box.png", f"/products/151c-{key}-slim.png", excel_img or ""]
            )
        else:
            cover = pick_image(
                [
                    f"/products/151c-{key}-{kind}.png",
                    f"/products/151c-{key}-box.png",
                    excel_img or "",
                ]
            )
        return f"{cover},{tops_151}" if tops_151 else cover

    if "151" in name and ("收藏家" in name or "コレクタ" in (excel_img or "")):
        cover = pick_image(["/products/151-box.png", excel_img or ""])
        return f"{cover},{tops_151}" if tops_151 else cover

    return pick_image([excel_img or "", "/products/placeholder.svg"])


def opc_code(series: str | None, name: str) -> str | None:
    m = re.search(r"OPC-(\d+)", f"{series or ''} {name}", re.I)
    return m.group(1).zfill(2) if m else None


def ebc_prb_stc_slug(name: str, box_type: str) -> str | None:
    hay = name
    m = re.search(r"(EBC|PRBC|STC|PRB)-?(\d+)", hay, re.I)
    if not m:
        return None
    kind, num = m.group(1).upper(), m.group(2).zfill(2)
    if kind == "EBC" and num == "02":
        base = "ebc-02-anime-25th-collection"
    elif kind == "EBC" and num == "03":
        base = "ebc-03-heroines-edition"
    elif kind == "EBC" and num == "04":
        base = "onepiece-ebc-04"
    elif kind == "EBC" and num == "01":
        base = "onepiece-ebc-01-memorial"
    elif kind.startswith("PRB"):
        base = f"onepiece-prbc-{num}" if num != "01" else "onepiece-prbc-01"
        if num == "02":
            base = "onepiece-prbc-02"
    elif kind == "STC":
        base = f"onepiece-stc-{num}"
    else:
        base = f"onepiece-{kind.lower()}-{num}"
    suffix = {"原盒": "box", "散包": "pack", "原箱": "case", "特别版": "box"}.get(box_type, "box")
    # 旧库 EBC-02 无后缀
    if base == "ebc-02-anime-25th-collection" and suffix == "box":
        return f"{base}-box"
    if base.startswith("onepiece-") and suffix == "box":
        return f"{base}-cn" if not base.endswith("-cn") else base
    return f"{base}-{suffix}" if suffix != "box" or "official" in base else f"{base}-box"


def is_section(name: str) -> bool:
    s = name.strip()
    if not s:
        return True
    if s.startswith("【") or s.startswith("※"):
        return True
    # 分区标题（无 SKU）
    if s in {
        "EB 特别补充包",
        "PRB 豪华补充包 / 典藏包",
        "特殊卡包",
        "基本卡组 / 预组",
    }:
        return True
    if re.fullmatch(r"30周年庆典", s):
        return True
    if "（简中）" in s or "BOX" in s or re.search(r"OPC-|EBC-|STC-|PRB|PRBC", s):
        return False
    if any(k in s for k in ("弹 ·", "扩充包", "主弹", "预组", "礼盒区", "宝石包区", "周边", "套装区", "专题 ·")):
        return True
    if s.startswith("第") and "弹" in s:
        return True
    return False


def parse_excel() -> list[dict]:
    wb = load_workbook(XLSX, data_only=True)
    out: list[dict] = []
    sort_i = 10

    # —— 宝可梦 ——
    ws = wb["宝可梦"]
    for r in range(4, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or is_section(str(name)):
            continue
        series = ws.cell(r, 2).value
        box = ws.cell(r, 3).value
        spec = ws.cell(r, 4).value
        stock = ws.cell(r, 6).value
        price = ws.cell(r, 7).value
        cost = ws.cell(r, 8).value
        lang = ws.cell(r, 9).value or "简中"
        preorder = ws.cell(r, 10).value
        featured = ws.cell(r, 11).value
        ship = ws.cell(r, 12).value
        img = ws.cell(r, 13).value
        ptype = ws.cell(r, 15).value
        note = ws.cell(r, 16).value
        slug_x = ws.cell(r, 18).value
        box_type = normalize_pokemon_box(str(box) if box else None, str(spec) if spec else None)
        if not box_type:
            continue
        # 礼盒等非展开行：规格散包/原箱且名称未标明时跳过空壳
        if box_type in ("礼盒", "专属礼盒", "大礼盒", "构筑套装", "对战套装", "周边") and spec in (
            "散包",
            "原箱",
        ):
            continue
        title = csv_title_from_name(str(name))
        slug = pokemon_slug(
            title,
            box_type,
            str(series) if series else None,
            str(slug_x) if slug_x else None,
            str(name),
        )
        images = pokemon_images(box_type, str(series) if series else None, str(img) if img else None, str(name))
        cat = "宝可梦原盒"
        if box_type in ("礼盒", "专属礼盒", "大礼盒"):
            cat = "宝可梦礼盒"
        elif box_type in ("周边",):
            cat = "宝可梦周边"
        elif box_type in ("构筑套装", "对战套装"):
            cat = "宝可梦套装"
        out.append(
            {
                "name": format_sku_name(str(name), box_type),
                "slug": slug,
                "series": str(series).strip() if series else None,
                "boxType": box_type,
                "category": cat,
                "language": str(lang),
                "priceJpy": int(price or 0),
                "stock": int(stock or 0),
                "costPrice": int(cost) if cost not in (None, "") else None,
                "images": images,
                "featured": 1 if featured in (True, "是", "YES", "yes", 1) else 0,
                "isPreorder": 1 if preorder in (True, "是", "YES", "yes", 1) or str(ws.cell(r, 5).value) == "预售" else 0,
                "shippingDays": int(ship or 6),
                "description": str(note).strip() if note else None,
                "catalogSort": sort_i,
                "game": "pokemon",
                "productType": str(ptype) if ptype else None,
            }
        )
        sort_i += 10

    # —— 海贼王 ——
    ws = wb["海贼王"]
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or is_section(str(name)):
            continue
        series = ws.cell(r, 2).value
        box = ws.cell(r, 3).value
        spec = ws.cell(r, 4).value
        stock = ws.cell(r, 6).value
        price = ws.cell(r, 7).value
        cost = ws.cell(r, 8).value
        lang = ws.cell(r, 9).value or "简中"
        preorder = ws.cell(r, 11).value
        featured = ws.cell(r, 12).value
        release = ws.cell(r, 13).value
        ship = ws.cell(r, 14).value
        img = ws.cell(r, 15).value
        note = ws.cell(r, 18).value
        slug_x = ws.cell(r, 20).value
        s_spec = str(spec).strip() if spec else "原盒"
        box_s = str(box).strip() if box else ""
        if box_s in ("礼盒", "专属礼盒", "大礼盒", "周边"):
            box_type = box_s
        else:
            box_type = s_spec if s_spec in ("原盒", "散包", "原箱") else (box_s or "原盒")
        if box_type == "特别版":
            box_type = "原盒"
        code = opc_code(str(series) if series else None, str(name))
        slug = None
        if code:
            base = OPC_SLUG_BASE.get(code, f"opc-{code}-official")
            suffix = {"原盒": "box", "散包": "pack", "原箱": "case"}[box_type]
            if base.endswith("-official"):
                slug = f"{base}-{suffix}-cn"
            else:
                slug = f"{base}-{suffix}" if suffix != "box" else f"{base}-box"
        if not slug:
            slug = ebc_prb_stc_slug(str(name), box_type)
        if not slug and slug_x:
            slug = str(slug_x).strip()
            if box_type == "散包" and slug.endswith("-box"):
                slug = slug[:-4] + "-pack"
            if box_type == "原箱" and slug.endswith("-box"):
                slug = slug[:-4] + "-case"
        if not slug:
            slug = re.sub(r"[^\w\u4e00-\u9fff-]+", "-", str(name))[:80]

        cover = None
        if code:
            cover = f"/products/onepiece-official/opc-{code}-pack.png"
        if img:
            cover = pick_image([str(img), cover or ""])
        elif cover:
            cover = pick_image([cover])
        else:
            # EBC/STC：名称或系列列均可识别
            hay = f"{name} {series or ''}"
            m = re.search(r"(EBC|PRBC|STC)-?(\d+)", hay, re.I)
            if m:
                k, n = m.group(1).lower(), m.group(2).zfill(2)
                if k == "ebc":
                    cover = pick_image(
                        [
                            f"/products/onepiece-official/ebc-{n}-pack.png",
                            f"/products/ebc-{n}-box.jpg",
                            f"/products/ebc-{n}-box.png",
                        ]
                    )
                elif k.startswith("prb"):
                    cover = pick_image([f"/products/onepiece-official/prbc-{n}-pack.png"])
                else:
                    cover = pick_image(
                        [
                            f"/products/onepiece-official/stc-{n}-deck.png",
                            f"/products/onepiece-official/stc-{n}-deck.jpg",
                        ]
                    )
        cover = cover or "/products/placeholder.png"

        series_s = str(series).strip() if series else None
        if code and (not series_s or series_s == f"OPC-{code}"):
            # 尽量用全称
            titles = {
                "12": "师徒之绊",
                "13": "继承的意志",
                "14": "苍海七杰",
                "15": "神之岛的冒险",
                "16": "决战之时",
            }
            if code in titles:
                series_s = f"OPC-{code} {titles[code]}"

        rel = None
        if release:
            try:
                if hasattr(release, "isoformat"):
                    rel = release.isoformat()
                else:
                    rel = str(release)
            except Exception:
                rel = None

        display = format_sku_name(str(name), box_type)
        for code_k, cn in EBC_DISPLAY.items():
            if code_k in display and "动画" not in display and "女英雄" not in display and "艾格" not in display:
                # 英文表名 → 中文展示名，保留规格后缀
                suffix = ""
                for s in ("原盒", "散包", "原箱"):
                    if display.endswith(f"{s}（简中）"):
                        suffix = f" {s}（简中）"
                        break
                if suffix:
                    display = f"{cn}{suffix}"
                break

        out.append(
            {
                "name": display,
                "slug": slug,
                "series": series_s,
                "boxType": box_type,
                "category": "海贼王",
                "language": str(lang),
                "priceJpy": int(price or 0),
                "stock": int(stock or 0),
                "costPrice": int(cost) if cost not in (None, "") else None,
                "images": cover,
                "featured": 1 if featured in (True, "是", 1) else 0,
                "isPreorder": 1 if preorder in (True, "是", 1) else 0,
                "shippingDays": int(ship or 6),
                "description": str(note).strip() if note else None,
                "catalogSort": sort_i,
                "releaseDate": rel,
                "game": "onepiece",
            }
        )
        sort_i += 10

    return out


def merge_topcards(conn: sqlite3.Connection, slug: str, cover: str) -> str:
    row = conn.execute("SELECT images FROM Product WHERE slug=?", (slug,)).fetchone()
    tops: list[str] = []
    if row and row[0]:
        parts = [p.strip() for p in str(row[0]).split(",") if p.strip()]
        tops = [p for p in parts[1:] if "/topcards/" in p]
    # manifest fallback by stem
    return ",".join([cover] + tops) if tops else cover


def ensure_catalog_sort_col(conn: sqlite3.Connection) -> None:
    cols = {r[1] for r in conn.execute("PRAGMA table_info(Product)")}
    if "catalogSort" not in cols:
        conn.execute('ALTER TABLE Product ADD COLUMN catalogSort INTEGER NOT NULL DEFAULT 999999')


def upsert(conn: sqlite3.Connection, items: list[dict]) -> tuple[int, int]:
    ensure_catalog_sort_col(conn)
    inserted = updated = 0
    for it in items:
        images = merge_topcards(conn, it["slug"], it["images"])
        existing = conn.execute("SELECT id FROM Product WHERE slug=?", (it["slug"],)).fetchone()
        ts = now()
        if existing:
            conn.execute(
                """
                UPDATE Product SET
                  name=?, series=?, category=?, language=?, description=?,
                  priceJpy=?, stock=?, images=?, boxType=?, featured=?,
                  isPreorder=?, shippingDays=?, status='上架',
                  costPrice=COALESCE(?, costPrice),
                  catalogSort=?, releaseDate=COALESCE(?, releaseDate),
                  updatedAt=?
                WHERE slug=?
                """,
                (
                    it["name"],
                    it["series"],
                    it["category"],
                    it["language"],
                    it.get("description"),
                    it["priceJpy"],
                    it["stock"],
                    images,
                    it["boxType"],
                    it["featured"],
                    it["isPreorder"],
                    it["shippingDays"],
                    it.get("costPrice"),
                    it["catalogSort"],
                    it.get("releaseDate"),
                    ts,
                    it["slug"],
                ),
            )
            updated += 1
        else:
            conn.execute(
                """
                INSERT INTO Product (
                  id, name, slug, category, series, language, description,
                  priceJpy, stock, images, boxType, featured, isPreorder,
                  shippingDays, releaseDate, costPrice, status, researchStatus,
                  catalogSort, createdAt, updatedAt
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    str(uuid.uuid4()),
                    it["name"],
                    it["slug"],
                    it["category"],
                    it["series"],
                    it["language"],
                    it.get("description"),
                    it["priceJpy"],
                    it["stock"],
                    images,
                    it["boxType"],
                    it["featured"],
                    it["isPreorder"],
                    it["shippingDays"],
                    it.get("releaseDate"),
                    it.get("costPrice"),
                    "上架",
                    "表格同步",
                    it["catalogSort"],
                    ts,
                    ts,
                ),
            )
            inserted += 1
    return inserted, updated


def delist_orphans(conn: sqlite3.Connection, items: list[dict]) -> int:
    """表格外的旧宝可梦/海贼王 SKU 下架，避免重复列表。"""
    slugs = {it["slug"] for it in items}
    if not slugs:
        return 0
    rows = conn.execute(
        """
        SELECT slug FROM Product
        WHERE status='上架'
          AND (
            category LIKE '宝可梦%'
            OR category='海贼王'
            OR category LIKE '%ONE%'
            OR series LIKE 'OPC%'
            OR series LIKE 'EB%'
            OR series LIKE 'EBC%'
            OR series LIKE 'PRB%'
            OR series LIKE 'STC%'
            OR series LIKE '航海王%'
            OR name LIKE '收集啦151%'
            OR slug LIKE 'pokemon-151%'
            OR slug LIKE '收集啦151%'
            OR slug LIKE 'onepiece-gift%'
          )
        """
    ).fetchall()
    orphan = [r[0] for r in rows if r[0] not in slugs]
    ts = now()
    for slug in orphan:
        conn.execute(
            "UPDATE Product SET status='下架', updatedAt=? WHERE slug=?",
            (ts, slug),
        )
    return len(orphan)


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"找不到表格: {XLSX}")
    print(f"使用表格: {XLSX}")
    items = parse_excel()
    # 同 slug 后者覆盖（表格后行优先）
    by_slug: dict[str, dict] = {}
    for it in items:
        by_slug[it["slug"]] = it
    items = sorted(by_slug.values(), key=lambda x: x["catalogSort"])
    print(f"解析表格 SKU: {len(items)}（宝可梦+海贼王，去重后）")
    print("前5:", [(i["catalogSort"], i["slug"], i["boxType"]) for i in items[:5]])
    print("后5:", [(i["catalogSort"], i["slug"], i["boxType"]) for i in items[-5:]])
    for db in DBS:
        if not db.exists():
            print("跳过缺失库", db)
            continue
        conn = sqlite3.connect(db)
        ins, upd = upsert(conn, items)
        off = delist_orphans(conn, items)
        conn.commit()
        n = conn.execute("SELECT COUNT(*) FROM Product WHERE status='上架'").fetchone()[0]
        conn.close()
        print(f"{db.name}: +{ins} ~{upd} 下架旧SKU={off} 上架总数={n}")


if __name__ == "__main__":
    main()
