# -*- coding: utf-8 -*-
"""对照 Excel 与网站 DB：未匹配商品、缺图、命名差异。"""
from __future__ import annotations

import re
import sqlite3
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\33092\Desktop\网站数据.xlsx")
DB = Path(r"C:\Users\33092\Documents\card-shop\dev.db")
PUBLIC = Path(r"C:\Users\33092\Documents\card-shop\public")
OUT = Path(r"C:\Users\33092\Documents\card-shop\scripts\_audit-excel-vs-site.txt")


def norm(s: object) -> str:
    t = str(s or "")
    t = t.replace("（", "(").replace("）", ")")
    t = re.sub(r"\s+", "", t)
    return t.lower()


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    db = sqlite3.connect(DB)
    db.row_factory = sqlite3.Row

    excel = []
    for sheet in ["宝可梦", "海贼王", "火影"]:
        ws = wb[sheet]
        for r in range(5, ws.max_row + 1):
            name = ws.cell(r, 1).value
            series = ws.cell(r, 2).value
            box = ws.cell(r, 3).value
            status = ws.cell(r, 4).value
            stock = ws.cell(r, 5).value
            price = ws.cell(r, 7).value
            if not name or not series:
                continue
            excel.append(
                {
                    "sheet": sheet,
                    "name": str(name).strip(),
                    "series": str(series).strip(),
                    "box": str(box).strip() if box else "",
                    "status": str(status).strip() if status else "",
                    "stock": stock,
                    "price": price,
                }
            )

    products = list(
        db.execute(
            "SELECT id,name,slug,series,boxType,status,stock,priceJpy,images,language FROM Product"
        )
    )
    db_by_name = {norm(p["name"]): p for p in products}
    db_by_slug = {p["slug"]: p for p in products}

    SLUG_ALIASES = {
        "强化包 极巨攻防 BOX（简中）": "极巨攻防-box-简中",
        "强化包 璀璨反击 BOX（简中）": "璀璨反击-box-简中",
        "强化包 怒炎灼天 BOX（简中）": "怒炎灼天-box-简中",
        "强化包 终末炎舞 BOX（简中）": "终末炎舞-box-简中",
        "强化包 暗影夺辉 BOX（简中）": "暗影夺辉-box-简中",
        "强化包 胜象星引 BOX（简中）": "胜象星引-box-简中",
        "强化包 第一弹 BOX（简中）": "sm-强化包-第一弹-box-简中",
        "强化包 第二弹 BOX（简中）": "sm-强化包-第二弹-box-简中",
        "补充包 强化扩张包 朱紫ex BOX（简中）": "朱-紫ex-強化拡張パック-box-简中",
        "补充包 奇迹启程 肥盒 BOX（简中）": "奇迹启程-box-简中",
        "补充包 无畏太晶 肥盒 BOX（简中）": "无畏太晶-box-简中",
        "补充包 嘉奖回合 肥盒 BOX（简中）": "嘉奖回合-box-简中",
        "补充包 黑晶炽诚 肥盒 BOX（简中）": "黑晶炽焰-box-简中",
        "补充包 真实玄虚 肥盒 BOX（简中）": "真实玄虚-box-简中",
        "补充包 利刃猛醒 肥盒 BOX（简中）": "剑刃觉醒-box-简中",
        "补充包 璀璨诡幻 肥盒 BOX（简中）": "璀璨诡幻-box-简中",
        "补充包 星彩晶璃 肥盒 BOX（简中）": "星彩晶璃-box-简中",
        "补充包 共逐荣光 肥盒 BOX（简中）": "共逐荣光-box-简中",
        "补充包 奇迹启程 瘦盒 BOX（简中）": "奇迹启程-slim-box-简中",
        "补充包 无畏太晶 瘦盒 BOX（简中）": "无畏太晶-slim-box-简中",
        "补充包 嘉奖回合 瘦盒 BOX（简中）": "嘉奖回合-slim-box-简中",
        "补充包 黑晶炽诚 瘦盒 BOX（简中）": "黑晶炽焰-slim-box-简中",
        "补充包 真实玄虚 瘦盒 BOX（简中）": "真实玄虚-slim-box-简中",
        "补充包 利刃猛醒 瘦盒 BOX（简中）": "剑刃觉醒-slim-box-简中",
        "补充包 璀璨诡幻 瘦盒 BOX（简中）": "璀璨诡幻-slim-box-简中",
        "补充包 星彩晶璃 瘦盒 BOX（简中）": "星彩晶璃-slim-box-简中",
        "补充包 共逐荣光 瘦盒 BOX（简中）": "共逐荣光-slim-box-简中",
    }

    missing = []
    present = []
    skipped_pending = []
    for e in excel:
        pending = ("官方名称" in e["name"]) or ("待官方" in e["name"]) or (e["price"] in (None, ""))
        n = norm(e["name"])
        alts = [n]
        if n.startswith("补充包"):
            alts.append(n[3:])
        if n.startswith("强化包"):
            alts.append(n[3:])
        hit = next((db_by_name[a] for a in alts if a in db_by_name), None)
        if not hit and e["name"] in SLUG_ALIASES:
            hit = db_by_slug.get(SLUG_ALIASES[e["name"]])
        if not hit:
            # series+boxtype loose match for primary boxes
            for p in products:
                if norm(p["series"]).startswith(norm(e["series"])[:12]) and e["name"][:6] in p["name"]:
                    hit = p
                    break
        if not hit:
            if pending:
                skipped_pending.append(e)
            else:
                missing.append(e)
        else:
            present.append((e, dict(hit)))

    broken = []
    for p in products:
        imgs = [x.strip() for x in (p["images"] or "").split(",") if x.strip()]
        if not imgs:
            broken.append((p["name"], p["slug"], "EMPTY"))
            continue
        first = imgs[0]
        if first.startswith("/"):
            fp = PUBLIC / first.lstrip("/")
            if not fp.exists():
                broken.append((p["name"], p["slug"], first))

    # 图片库未匹配
    img_ws = wb["图片库"]
    unmatched_imgs = []
    for r in range(2, img_ws.max_row + 1):
        if img_ws.cell(r, 4).value == "未匹配":
            unmatched_imgs.append(
                (img_ws.cell(r, 2).value, img_ws.cell(r, 3).value, img_ws.cell(r, 9).value)
            )

    # DB 上架但 Excel 无对应（主规格）
    excel_names = {norm(e["name"]) for e in excel}
    orphans = []
    for p in products:
        if p["boxType"] in ("散包", "原箱", "肥散包", "瘦散包", "肥原箱", "瘦原箱"):
            continue
        if norm(p["name"]) not in excel_names and not any(
            norm(p["name"]).startswith(norm(e["name"])[:8]) for e in excel if e["name"]
        ):
            orphans.append(p)

    lines = []
    lines.append(f"Excel 商品行: {len(excel)}")
    lines.append(f"DB 商品: {len(products)}")
    lines.append(f"Excel 有价但网站未匹配: {len(missing)}")
    for e in missing:
        lines.append(
            f"  MISSING [{e['sheet']}] {e['name']} | {e['series']} | {e['box']} | {e['status']} | price={e['price']} stock={e['stock']}"
        )
    lines.append(f"\nExcel 待官方/无价跳过: {len(skipped_pending)}")
    for e in skipped_pending[:30]:
        lines.append(f"  PENDING {e['name']} | {e['series']}")
    lines.append(f"\n缺本地图片文件: {len(broken)}")
    for b in broken:
        lines.append(f"  NOFILE {b[0]} | {b[1]} | {b[2]}")
    lines.append(f"\n图片库未匹配: {len(unmatched_imgs)}")
    for u in unmatched_imgs:
        lines.append(f"  IMG {u[0]} | {u[1]} | {u[2]}")
    lines.append(f"\n网站有但 Excel 主规格难匹配: {len(orphans)}")
    for p in orphans[:40]:
        lines.append(f"  ORPHAN {p['name']} | {p['series']} | {p['boxType']}")

    # 售罄与未上架 sheet
    sold = wb["售罄与未上架"]
    lines.append("\n售罄与未上架表:")
    for r in range(2, sold.max_row + 1):
        lines.append(
            f"  {sold.cell(r,1).value} | {sold.cell(r,2).value} | 网站={sold.cell(r,5).value} | stock={sold.cell(r,6).value}"
        )

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(OUT)
    print("\n".join(lines[:80]))
    print("...")
    print(f"total missing={len(missing)} broken_imgs={len(broken)} unmatched_img_lib={len(unmatched_imgs)}")


if __name__ == "__main__":
    main()
