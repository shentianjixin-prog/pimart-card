# -*- coding: utf-8 -*-
"""补缺图、同步 Excel 库存价到已有商品、检查翻译完整性。"""
from __future__ import annotations

import re
import shutil
import sqlite3
import ssl
import urllib.request
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public" / "products"
DBS = [ROOT / "dev.db", ROOT / "prisma" / "data" / "initial.db"]
XLSX = Path(r"c:\Users\33092\Desktop\网站数据.xlsx")
TRANS = ROOT / "src" / "lib" / "translations.ts"

CTX = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode = ssl.CERT_NONE

# 缺文件的商品：优先用已有相似图 / 官网图
IMAGE_FIXES = {
    "30th-deluxe-box.png": [
        "https://image.pokemon.com.cn/wp-content/uploads/2026/02/260220_30th_CGWL_650_488.png",
        "30th-box.png",
        "csv10c-box.png",
    ],
    "lillie-battle-box.png": [
        "https://image.pokemon.com.cn/wp-content/uploads/2026/05/260515_CSV10_LLY_650_488.png",
        "csv10c-box.png",
    ],
    "n-battle-box.png": [
        "https://image.pokemon.com.cn/wp-content/uploads/2026/05/260515_CSV10_N_650_488.png",
        "csv10c-box.png",
    ],
}

# Excel 名称 → 网站现有 slug（命名对齐）
NAME_ALIASES = {
    "强化包 极巨攻防 BOX（简中）": "极巨攻防-box-简中",
    "强化包 璀璨反击 BOX（简中）": "璀璨反击-box-简中",
    "强化包 怒炎灼天 BOX（简中）": "怒炎灼天-box-简中",
    "强化包 终末炎舞 BOX（简中）": "终末炎舞-box-简中",
    "强化包 暗影夺辉 BOX（简中）": "暗影夺辉-box-简中",
    "强化包 胜象星引 BOX（简中）": "胜象星引-box-简中",
    "补充包 强化扩张包 朱紫ex BOX（简中）": None,  # 查现有
}


def download(url: str, dest: Path) -> bool:
    try:
        req = urllib.request.Request(
            url,
            headers={"User-Agent": "Mozilla/5.0", "Referer": "https://www.pokemon.cn/"},
        )
        with urllib.request.urlopen(req, timeout=45, context=CTX) as resp:
            data = resp.read()
        if data[:1] == b"<" or len(data) < 1000:
            print(f"  skip bad download {dest.name}")
            return False
        dest.write_bytes(data)
        print(f"  downloaded {dest.name} ({len(data)} bytes)")
        return True
    except Exception as e:
        print(f"  download fail {dest.name}: {e}")
        return False


def ensure_image(filename: str) -> bool:
    dest = PUBLIC / filename
    if dest.is_file() and dest.stat().st_size > 1000:
        return True
    sources = IMAGE_FIXES.get(filename, [])
    for src in sources:
        if src.startswith("http"):
            if download(src, dest):
                return True
        else:
            cand = PUBLIC / src
            if cand.is_file():
                shutil.copy2(cand, dest)
                print(f"  copied {src} -> {filename}")
                return True
    return dest.is_file()


def check_translations() -> list[str]:
    text = TRANS.read_text(encoding="utf-8")
    issues = []
    # match object literals key: { zh: "...", ja: "...", en: "..." }
    for m in re.finditer(
        r"^\s*([A-Za-z0-9_]+):\s*\{([^\n]*zh:[^\n]*)\}",
        text,
        re.M,
    ):
        key, body = m.group(1), m.group(2)
        for lang in ("zh", "ja", "en"):
            if f"{lang}:" not in body:
                issues.append(f"{key} missing {lang}")
                continue
            vm = re.search(rf'{lang}:\s*"([^"]*)"', body)
            if vm is not None and vm.group(1).strip() == "":
                issues.append(f"{key} empty {lang}")
    return issues


def sync_excel_prices(db_path: Path) -> None:
    if not db_path.exists():
        return
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    db = sqlite3.connect(db_path)
    db.row_factory = sqlite3.Row
    updated = 0

    # build name->row map from excel priced products
    rows = []
    for sheet in ("宝可梦", "海贼王"):
        ws = wb[sheet]
        for r in range(5, ws.max_row + 1):
            name = ws.cell(r, 1).value
            series = ws.cell(r, 2).value
            box = ws.cell(r, 3).value
            status = ws.cell(r, 4).value
            stock = ws.cell(r, 5).value
            price = ws.cell(r, 7).value
            if not name or not series or price in (None, ""):
                continue
            if "官方名称" in str(name) or "待官方" in str(name):
                continue
            rows.append((str(name).strip(), str(series).strip(), str(box or "").strip(), status, stock, float(price)))

    # slug aliases
    alias_slug = {
        "强化包 极巨攻防 BOX（简中）": "极巨攻防-box-简中",
        "强化包 璀璨反击 BOX（简中）": "璀璨反击-box-简中",
        "强化包 怒炎灼天 BOX（简中）": "怒炎灼天-box-简中",
        "强化包 终末炎舞 BOX（简中）": "终末炎舞-box-简中",
        "强化包 暗影夺辉 BOX（简中）": "暗影夺辉-box-简中",
        "强化包 胜象星引 BOX（简中）": "胜象星引-box-简中",
    }

    find = db.prepare if False else None
    for name, series, box, status, stock, price in rows:
        # find product
        p = db.execute("SELECT id,name,slug,stock,priceJpy,status FROM Product WHERE name=?", (name,)).fetchone()
        if not p and name in alias_slug:
            p = db.execute(
                "SELECT id,name,slug,stock,priceJpy,status FROM Product WHERE slug=?",
                (alias_slug[name],),
            ).fetchone()
        if not p:
            # try strip 补充包
            alt = name.replace("补充包 ", "").replace("强化包 ", "")
            p = db.execute("SELECT id,name,slug,stock,priceJpy,status FROM Product WHERE name=?", (alt,)).fetchone()
        if not p:
            continue
        new_stock = int(stock or 0)
        # 网站库存策略：Excel 售罄=0，在售用 Excel；若 Excel 库存>0 则同步
        # 用户要求全部上架
        db.execute(
            "UPDATE Product SET priceJpy=?, stock=?, status='上架', updatedAt=? WHERE id=?",
            (int(price), new_stock, __import__("datetime").datetime.utcnow().isoformat() + "Z", p["id"]),
        )
        updated += 1

    # ensure all products 上架
    db.execute("UPDATE Product SET status='上架' WHERE status IS NULL OR status!=?", ("上架",))
    db.commit()
    print(f"[{db_path.name}] synced excel rows touched≈{updated}, all status=上架")
    db.close()


def list_products_without_images() -> None:
    db = sqlite3.connect(DBS[0])
    for r in db.execute("SELECT name,slug,images FROM Product"):
        imgs = [x.strip() for x in (r[2] or "").split(",") if x.strip()]
        if not imgs:
            print("EMPTY", r[0])
            continue
        fp = ROOT / "public" / imgs[0].lstrip("/")
        if not fp.exists():
            print("MISSING FILE", r[0], imgs[0])


def main() -> None:
    print("=== 翻译检查 ===")
    issues = check_translations()
    print(f"issues: {len(issues)}")
    for i in issues[:20]:
        print(" ", i)

    print("\n=== 补缺图 ===")
    for fn in IMAGE_FIXES:
        ok = ensure_image(fn)
        print(fn, "OK" if ok else "FAIL")

    print("\n=== 同步 Excel 价格库存 + 全部上架 ===")
    for dbp in DBS:
        sync_excel_prices(dbp)

    print("\n=== 缺图复查 ===")
    list_products_without_images()


if __name__ == "__main__":
    main()
