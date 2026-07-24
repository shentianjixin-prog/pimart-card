# -*- coding: utf-8 -*-
"""
按网站数据表收尾：
1) 替换 placeholder，肥瘦盒封面：瘦规格用 *-slim（无则 box），肥规格用 *-fat（无则 box）
2) 导出 prisma/data/website-catalog.json 供 Railway
3) 回写 Excel「图片」列 +「图片库」备份（桌面 + prisma/data）
"""
from __future__ import annotations

import json
import re
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public" / "products"
DBS = [ROOT / "dev.db", ROOT / "prisma" / "data" / "initial.db"]
XLSX_DESKTOP = Path(r"c:\Users\33092\Desktop\网站数据.xlsx")
XLSX_REPO = ROOT / "prisma" / "data" / "网站数据.xlsx"
CATALOG = ROOT / "prisma" / "data" / "website-catalog.json"

# 系列代码 → 图片文件主干（不含 -slim/-fat/-box）
SERIES_STEM = {
    "CSV10c": "csv10c",
    "CSV9.5c": "csv9c",
    "CSV9c": "csv9c",
    "CSV8c": "csv8c",
    "CSV7c": "csv7c",
    "CSV6c": "csv6c",
    "CSV5c": "csv5c",
    "CSV4c": "csv4c",
    "CSV3c": "csv3c",
    "CSV2c": "csv2c",
    "CSV1c": "csv1c",
    "CS6b": "cs6b",
    "CS6a": "cs6a",
    "CS5b": "cs5b",
    "CS5a": "cs5a",
    "CS4b": "cs4b",
    "CS4a": "cs4a",
    "CS3b": "cs3b",
    "CS3a": "cs3a",
    "CS2b": "cs2b",
    "CS2a": "cs2a",
    "CS1b": "cs1b",
    "CS1a": "cs1a",
    "CS1.5": "cs15",
    "CS2.5": "cs25",
    "CS3.5": "cs35",
    "CS4.5": "cs45",
    "CS5.5": "cs55",
    "CS6.5": "cs65",
    "CSM1a": "csm1a",
    "CSM1b": "csm1b",
    "CSM1c": "csm1c",
    "CSM2a": "csm2a",
    "CSM2b": "csm2b",
    "CSM2c": "csm2c",
}

NAME_STEM = [
    (r"碧海暗影\s*逐", "cs6b"),
    (r"碧海暗影\s*啸", "cs6a"),
    (r"碧海暗影", "cs6a"),
    (r"勇魅群星\s*勇", "cs5b"),
    (r"勇魅群星\s*魅", "cs5a"),
    (r"九彩汇聚\s*朋", "cs4a"),
    (r"九彩汇聚\s*源", "cs4b"),
    (r"洪荒演武\s*激", "cs3b"),
    (r"洪荒演武\s*茂", "cs3a"),
    (r"浓墨重彩\s*靛", "cs2b"),
    (r"浓墨重彩\s*黎", "cs2a"),
    (r"极巨争锋\s*焰", "cs1b"),
    (r"极巨争锋\s*雷", "cs1a"),
    (r"交相辉映\s*唤", "csm2c"),
    (r"交相辉映\s*沐", "csm2a"),
    (r"交相辉映\s*魁", "csm2b"),
    (r"横空出世\s*泽", "csm1c"),
    (r"横空出世\s*苍", "csm1b"),
    (r"横空出世\s*赫", "csm1a"),
    (r"极巨攻防", "cs15"),
    (r"璀璨反击", "cs25"),
    (r"怒炎灼天", "cs35"),
    (r"终末炎舞", "cs45"),
    (r"暗影夺辉", "cs55"),
    (r"胜象星引", "cs65"),
]


def exists(rel: str) -> bool:
    return (ROOT / "public" / rel.lstrip("/")).exists()


def pick(*cands: str) -> str | None:
    for c in cands:
        if c and exists(c):
            return c
    return None


def stem_for(series: str | None, name: str) -> str | None:
    s = series or ""
    for code, stem in SERIES_STEM.items():
        if code.lower() in s.lower() or code in s:
            return stem
    for pat, stem in NAME_STEM:
        if re.search(pat, name):
            return stem
    m = re.search(r"CSV\s*(\d+(?:\.\d+)?)c", s, re.I)
    if m:
        return f"csv{m.group(1).replace('.', '')}c".replace("csv95c", "csv9c")
    m = re.search(r"CSM?\s*(\d+[a-z]?)", s, re.I)
    if m:
        return f"cs{m.group(1).lower()}" if "CSM" not in s.upper() else f"csm{m.group(1).lower()}"
    return None


def cover_for(box_type: str, stem: str | None, current: str) -> str:
    """瘦规格优先 slim；肥规格优先 fat；都没有则 box。"""
    tops = []
    if current:
        parts = [p.strip() for p in current.split(",") if p.strip()]
        cover0 = parts[0] if parts else ""
        tops = [p for p in parts[1:] if p.startswith("/products/topcards/")]
    else:
        cover0 = ""

    if cover0 and "placeholder" not in cover0 and exists(cover0):
        # 已有有效封面：若是肥规格却挂了 slim，或瘦规格挂了 fat，则纠正
        prefer_slim = box_type.startswith("瘦") or box_type in ("瘦盒", "瘦散包", "瘦原箱")
        prefer_fat = box_type.startswith("肥") or box_type in ("肥盒", "肥散包", "肥原箱")
        if stem:
            slim = f"/products/{stem}-slim.png"
            fat = f"/products/{stem}-fat.png"
            box = pick(f"/products/{stem}-box.png", f"/products/{stem}-box.jpg", f"/products/{stem}-box.webp")
            if prefer_slim:
                chosen = pick(slim, box, cover0, fat)
            elif prefer_fat:
                chosen = pick(fat, box, cover0, slim)
            else:
                # 默认瘦盒封面
                chosen = pick(slim, box, cover0, fat)
            if chosen:
                return ",".join([chosen] + tops) if tops else chosen
        return current

    if not stem:
        return current if current and "placeholder" not in current else "/products/placeholder.svg"

    slim = f"/products/{stem}-slim.png"
    fat = f"/products/{stem}-fat.png"
    box = pick(f"/products/{stem}-box.png", f"/products/{stem}-box.jpg", f"/products/{stem}-box.webp")
    prefer_slim = box_type.startswith("瘦") or not box_type.startswith("肥")
    if prefer_slim:
        chosen = pick(slim, box, fat)
    else:
        chosen = pick(fat, box, slim)
    if not chosen:
        return current or "/products/placeholder.svg"
    return ",".join([chosen] + tops) if tops else chosen


def fix_db_images() -> dict:
    stats = {}
    for db_path in DBS:
        if not db_path.exists():
            continue
        con = sqlite3.connect(db_path)
        rows = con.execute("SELECT id, name, series, boxType, images, status FROM Product").fetchall()
        n = 0
        for pid, name, series, box_type, images, status in rows:
            stem = stem_for(series, name or "")
            # 151 四弹
            m151 = re.search(r"收集啦151\s*(旅|望|惊|聚)", f"{series or ''} {name or ''}")
            if m151:
                key = {"旅": "lv", "望": "wang", "惊": "jing", "聚": "ju"}[m151.group(1)]
                stem = f"151c-{key}"
            new_img = cover_for(box_type or "", stem, images or "")
            if new_img != (images or ""):
                con.execute("UPDATE Product SET images=?, updatedAt=CURRENT_TIMESTAMP WHERE id=?", (new_img, pid))
                n += 1
        con.commit()
        listed = con.execute("SELECT COUNT(*) FROM Product WHERE status='上架'").fetchone()[0]
        ph = con.execute(
            "SELECT COUNT(*) FROM Product WHERE status='上架' AND images LIKE '%placeholder%'"
        ).fetchone()[0]
        con.close()
        stats[db_path.name] = {"images_updated": n, "listed": listed, "placeholder_listed": ph}
    return stats


def export_catalog() -> int:
    db_path = ROOT / "dev.db"
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    cols = [r[1] for r in con.execute("PRAGMA table_info(Product)").fetchall()]
    rows = con.execute("SELECT * FROM Product WHERE status='上架' ORDER BY catalogSort, name").fetchall()
    items = []
    for r in rows:
        d = {k: r[k] for k in cols}
        # JSON 友好
        for k, v in list(d.items()):
            if hasattr(v, "isoformat"):
                d[k] = v.isoformat()
        items.append(d)
    con.close()
    CATALOG.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    return len(items)


def backup_excel_images() -> dict:
    src = XLSX_DESKTOP if XLSX_DESKTOP.exists() else XLSX_REPO
    wb = load_workbook(src)
    con = sqlite3.connect(ROOT / "dev.db")
    con.row_factory = sqlite3.Row
    by_slug = {
        r["slug"]: r
        for r in con.execute("SELECT slug, name, series, images, status FROM Product").fetchall()
    }
    # 也建 name 索引
    by_name = {}
    for r in by_slug.values():
        by_name.setdefault(r["name"], r)

    updated_cells = 0
    img_lib_rows = []

    def sync_sheet(ws, header_row: int, name_col=1, slug_col=18, img_col=13):
        nonlocal updated_cells
        # detect columns from header
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v:
                headers[str(v).strip()] = c
        name_col = headers.get("商品名", name_col)
        slug_col = headers.get("SKU/Slug", slug_col)
        img_col = headers.get("图片", img_col)
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, name_col).value
            if not name:
                continue
            slug = ws.cell(r, slug_col).value if slug_col else None
            prod = None
            if slug and str(slug).strip() in by_slug:
                prod = by_slug[str(slug).strip()]
            elif str(name).strip() in by_name:
                prod = by_name[str(name).strip()]
            if not prod or not prod["images"]:
                continue
            cover = str(prod["images"]).split(",")[0].strip()
            if not cover.startswith("/"):
                continue
            old = ws.cell(r, img_col).value
            if old != cover:
                ws.cell(r, img_col, cover)
                updated_cells += 1
            img_lib_rows.append(
                {
                    "row": r,
                    "name": str(name).strip(),
                    "series": prod["series"] or "",
                    "status": "已备份" if exists(cover) else "缺文件",
                    "file": Path(cover).name,
                    "path": cover,
                    "note": f"DB status={prod['status']}",
                }
            )

    if "宝可梦" in wb.sheetnames:
        sync_sheet(wb["宝可梦"], 3, img_col=13, slug_col=18)
    if "海贼王" in wb.sheetnames:
        # 海贼王 header row 4, 图片 col 15, slug 20
        sync_sheet(wb["海贼王"], 4, img_col=15, slug_col=20)

    # 重建图片库
    if "图片库" in wb.sheetnames:
        ws = wb["图片库"]
        # 清内容保留表头
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        ws = wb.create_sheet("图片库")
        ws.append(["总表行号", "商品名", "系列/编号", "匹配状态", "标题", "本地文件名", "图片相对路径(网站)", "来源URL", "备注"])

    # 去重 path
    seen = set()
    for i, it in enumerate(img_lib_rows, start=2):
        key = (it["name"], it["path"])
        if key in seen:
            continue
        seen.add(key)
        ws.append(
            [
                it["row"],
                it["name"],
                it["series"],
                it["status"],
                it["name"],
                it["file"],
                it["path"],
                "",
                it["note"],
            ]
        )

    # 保存：桌面文件可能被 Excel 占用，失败则写旁路文件
    saved = None
    for target in (
        XLSX_DESKTOP,
        XLSX_DESKTOP.with_name("网站数据_已备份图片.xlsx"),
        XLSX_REPO,
    ):
        try:
            wb.save(target)
            saved = target
            break
        except PermissionError:
            continue
    if saved is None:
        raise PermissionError("无法写入 Excel，请先关闭桌面上的「网站数据.xlsx」后重试")
    if saved != XLSX_REPO:
        try:
            XLSX_REPO.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(saved, XLSX_REPO)
        except Exception as e:
            print("copy to repo failed:", e)
    if saved != XLSX_DESKTOP and XLSX_DESKTOP.exists():
        # 也尝试覆盖桌面原文件失败时保留旁路
        pass
    con.close()
    return {
        "excel_image_cells": updated_cells,
        "image_lib_unique": len(seen),
        "saved_as": str(saved),
        "repo_copy": str(XLSX_REPO),
    }


def main():
    print("fix images...", fix_db_images())
    n = export_catalog()
    print(f"catalog exported: {n} -> {CATALOG}")
    print("excel backup...", backup_excel_images())
    # 再统计 placeholder
    con = sqlite3.connect(ROOT / "dev.db")
    ph = con.execute(
        "SELECT name, boxType, images FROM Product WHERE status='上架' AND images LIKE '%placeholder%' LIMIT 20"
    ).fetchall()
    print("remaining placeholder listed:", len(ph))
    for r in ph:
        print(" ", r)
    print(
        "listed slim covers sample:",
        con.execute(
            "SELECT name, boxType, substr(images,1,50) FROM Product WHERE status='上架' AND boxType='瘦盒' LIMIT 8"
        ).fetchall(),
    )
    con.close()


if __name__ == "__main__":
    main()
